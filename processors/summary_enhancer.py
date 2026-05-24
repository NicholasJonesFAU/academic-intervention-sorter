"""
summary_enhancer.py — Enhances the Summary tab in output workbooks with
group breakdown charts, risk distribution, and contact coverage visuals.

Called by the Exporter after the basic summary rows are written.
"""

import io
import logging
from typing import Dict, Any, List

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

from utils.config import UNMATCHED_LOW_TAB, UNMATCHED_HIGH_TAB
from utils.excel_utils import _argb

logger = logging.getLogger("intervention_sorter")

COLOR_BG      = "#F4F6FB"
COLOR_PRIMARY = "#1F3864"
COLOR_ACCENT  = "#2F5496"
COLOR_GREEN   = "#2E7D32"
COLOR_RED     = "#C62828"
COLOR_AMBER   = "#F57F17"
COLOR_TEAL    = "#00695C"

CHART_W = 400
CHART_H = 260


class SummaryEnhancer:
    """Adds charts and visual summaries to the Summary worksheet."""

    def enhance(
        self,
        ws: Worksheet,
        metrics: Dict[str, Any],
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        start_row: int = None,
    ) -> None:
        """
        Add charts below the existing summary text.
        start_row: row to begin adding visuals (auto-detected if None)
        """
        if start_row is None:
            start_row = ws.max_row + 3

        # Chart 1 — Students by group (bar)
        group_counts = {
            tab: len(group_data.get(tab, pd.DataFrame()))
            for tab in group_order
        }
        group_counts[UNMATCHED_LOW_TAB]  = len(group_data.get(UNMATCHED_LOW_TAB, pd.DataFrame()))
        group_counts[UNMATCHED_HIGH_TAB] = len(group_data.get(UNMATCHED_HIGH_TAB, pd.DataFrame()))
        group_counts = {k: v for k, v in group_counts.items() if v > 0}

        if group_counts:
            img = self._make_group_bar(group_counts)
            ws.add_image(img, f"A{start_row}")
            start_row += 18

        # Chart 2 — Contact coverage donut
        matched  = metrics.get("contact_matches", 0)
        missed   = metrics.get("contact_misses", 0)
        if matched + missed > 0:
            img2 = self._make_contact_donut(matched, missed)
            ws.add_image(img2, f"A{start_row}")
            start_row += 14

        # Chart 3 — Risk course count distribution
        all_students = pd.concat(
            [df for df in group_data.values() if not df.empty],
            ignore_index=True,
        )
        if not all_students.empty and "Risk Course Count" in all_students.columns:
            img3 = self._make_risk_distribution(all_students)
            ws.add_image(img3, f"A{start_row}")

    # ------------------------------------------------------------------
    # Chart builders
    # ------------------------------------------------------------------

    def _make_group_bar(self, group_counts: Dict[str, int]) -> XLImage:
        labels = list(group_counts.keys())
        counts = list(group_counts.values())

        colors = []
        for label in labels:
            if label == UNMATCHED_HIGH_TAB:
                colors.append(COLOR_RED)
            elif label == UNMATCHED_LOW_TAB:
                colors.append(COLOR_AMBER)
            else:
                colors.append(COLOR_ACCENT)

        fig, ax = plt.subplots(figsize=(5.5, max(2.5, len(labels) * 0.38 + 0.8)),
                               facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)
        bars = ax.barh(labels, counts, color=colors, edgecolor="white")
        for bar, count in zip(bars, counts):
            ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                    str(count), va="center", fontsize=9,
                    fontweight="bold", color=COLOR_PRIMARY)
        ax.set_xlabel("Students", fontsize=9)
        ax.set_title("Students by Group", fontsize=11,
                     fontweight="bold", color=COLOR_PRIMARY)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.invert_yaxis()
        plt.tight_layout()
        return self._to_img(fig, CHART_W, CHART_H)

    def _make_contact_donut(self, matched: int, missed: int) -> XLImage:
        total = matched + missed
        pct = matched / total * 100 if total else 0

        fig, ax = plt.subplots(figsize=(3.5, 2.8), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)
        sizes = [matched, missed] if missed > 0 else [matched, 0.001]
        ax.pie(sizes, colors=[COLOR_GREEN, "#E0E0E0"], startangle=90,
               wedgeprops=dict(width=0.45, edgecolor="white", linewidth=1.5))
        ax.text(0, 0.1, f"{pct:.0f}%", ha="center", va="center",
                fontsize=16, fontweight="bold", color=COLOR_PRIMARY)
        ax.text(0, -0.2, "with contact", ha="center", va="center",
                fontsize=8, color="#546E7A")
        ax.set_title("Contact Coverage", fontsize=11,
                     fontweight="bold", color=COLOR_PRIMARY, pad=6)
        import matplotlib.patches as mpatches
        legend = [
            mpatches.Patch(color=COLOR_GREEN, label=f"Matched ({matched:,})"),
            mpatches.Patch(color="#E0E0E0",   label=f"No contact ({missed:,})"),
        ]
        ax.legend(handles=legend, loc="lower center",
                  bbox_to_anchor=(0.5, -0.08), ncol=2, fontsize=8, framealpha=0)
        plt.tight_layout()
        return self._to_img(fig, 300, 240)

    def _make_risk_distribution(self, df: pd.DataFrame) -> XLImage:
        counts = df["Risk Course Count"].value_counts().sort_index()
        labels = [f"{i} course{'s' if i != 1 else ''}" for i in counts.index]
        values = counts.values.tolist()

        colors = [COLOR_AMBER if i < 3 else COLOR_RED for i in counts.index]

        fig, ax = plt.subplots(figsize=(5, 2.8), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)
        bars = ax.bar(labels, values, color=colors, edgecolor="white")
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                    str(val), ha="center", va="bottom", fontsize=9,
                    fontweight="bold", color=COLOR_PRIMARY)
        ax.set_ylabel("Students", fontsize=9)
        ax.set_title("Risk Course Count Distribution", fontsize=11,
                     fontweight="bold", color=COLOR_PRIMARY)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout()
        return self._to_img(fig, CHART_W, 220)

    def _to_img(self, fig, width: int, height: int) -> XLImage:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=96, bbox_inches="tight",
                    facecolor=COLOR_BG)
        buf.seek(0)
        plt.close(fig)
        img = XLImage(buf)
        img.width  = width
        img.height = height
        return img
