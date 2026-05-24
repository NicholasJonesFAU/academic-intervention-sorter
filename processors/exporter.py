"""
exporter.py — Writes the final output Excel workbook.

Tab order:
  Summary → <Group Tabs> → Risk_1_2 → Risk_3_Plus → QA_Log → Processing_Manifest

All tabs use the standardized OUTPUT_COLUMNS schema.
"""

import logging
import platform
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.config import (
    APP_NAME,
    APP_VERSION,
    OUTPUT_COLUMNS,
    SUMMARY_TAB,
    QA_LOG_TAB,
    MANIFEST_TAB,
    UNMATCHED_LOW_TAB,
    UNMATCHED_HIGH_TAB,
    QA_LOG_COLUMNS,
    SUMMARY_LABELS,
    STYLE,
)
from utils.normalization import safe_excel_tab_name
from utils.excel_utils import (
    apply_data_tab_formatting,
    apply_summary_formatting,
    apply_qa_formatting,
    apply_manifest_formatting,
    make_header_fill,
    make_header_font,
    make_wrap_alignment,
)
from utils.logging_utils import QALog
from processors.summary_enhancer import SummaryEnhancer

logger = logging.getLogger("intervention_sorter")


class Exporter:
    """
    Generates the final output workbook from processed data.
    """

    def export(
        self,
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],           # safe tab names in processing order
        qa_log: QALog,
        metrics: Dict[str, Any],
        output_path: Path,
        source_files: Dict[str, str],
    ) -> None:
        """
        Build and save the output workbook.

        Args:
            group_data:   dict mapping safe_tab_name → DataFrame
            group_order:  ordered list of group tab names (excl. unmatched/QA/manifest)
            qa_log:       QALog instance with all QA events
            metrics:      dict of processing metrics for summary/manifest
            output_path:  resolved output .xlsx path
            source_files: dict mapping label → filename for manifest
        """
        logger.info("Exporter: Building output workbook → '%s'", output_path.name)

        wb = Workbook()
        wb.remove(wb.active)   # Remove default empty sheet
        used_tab_names: List[str] = []

        # 1. Summary tab
        self._write_summary(wb, metrics, source_files, group_data, group_order, used_tab_names)

        # 2. Group tabs (in order)
        for tab_name in group_order:
            df = group_data.get(tab_name, pd.DataFrame())
            self._write_data_tab(wb, tab_name, df, used_tab_names)

        # 3. Unmatched buckets
        for bucket_name in [UNMATCHED_LOW_TAB, UNMATCHED_HIGH_TAB]:
            df = group_data.get(bucket_name, pd.DataFrame())
            safe = safe_excel_tab_name(bucket_name, used_tab_names)
            used_tab_names.append(safe)
            self._write_data_tab(wb, safe, df, [])  # already registered

        # 4. Missing Contacts
        self._write_missing_contacts(wb, group_data, group_order, used_tab_names)

        # 5. QA_Log
        self._write_qa_log(wb, qa_log, used_tab_names)

        # 5. Processing_Manifest
        self._write_manifest(wb, metrics, source_files, used_tab_names)

        # Save
        try:
            wb.save(output_path)
            logger.info("Exporter: Workbook saved successfully → '%s'", output_path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Cannot save workbook. The file may be open in Excel.\n"
                f"Path: {output_path}\nError: {exc}"
            ) from exc

    # ------------------------------------------------------------------
    # Data tabs
    # ------------------------------------------------------------------

    def _write_data_tab(
        self,
        wb: Workbook,
        tab_name: str,
        df: pd.DataFrame,
        used_tab_names: List[str],
    ) -> None:
        """Write a standard data tab with OUTPUT_COLUMNS schema."""
        ws = wb.create_sheet(title=tab_name)
        used_tab_names.append(tab_name)

        # Ensure all output columns exist
        for col in OUTPUT_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        df_out = df[OUTPUT_COLUMNS].copy()
        df_out = df_out.fillna("")

        # Write header
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")

        # Apply formatting
        apply_data_tab_formatting(ws, OUTPUT_COLUMNS)

        logger.info(
            "Exporter: Tab '%s' written — %d rows.", tab_name, len(df_out)
        )

    # ------------------------------------------------------------------
    # Summary tab
    # ------------------------------------------------------------------

    def _write_summary(
        self,
        wb: Workbook,
        metrics: Dict[str, Any],
        source_files: Dict[str, str],
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        used_tab_names: List[str],
    ) -> None:
        safe = safe_excel_tab_name(SUMMARY_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        rows = []
        rows.append(("◆  PROCESSING SUMMARY  ◆", ""))
        rows.append(("", ""))
        rows.append(("Application", f"{APP_NAME} v{APP_VERSION}"))
        rows.append(("Processing Timestamp", metrics.get("processing_timestamp", "")))
        rows.append(("Execution Duration", f"{metrics.get('execution_duration', 0):.2f} seconds"))
        rows.append(("", ""))

        rows.append(("◆  INPUT FILE METRICS  ◆", ""))
        rows.append((SUMMARY_LABELS["total_input_rows"], metrics.get("total_input_rows", 0)))
        rows.append((SUMMARY_LABELS["total_at_risk_rows"], metrics.get("total_at_risk_rows", 0)))
        rows.append((SUMMARY_LABELS["duplicate_course_rows_removed"], metrics.get("duplicate_course_rows_removed", 0)))
        rows.append((SUMMARY_LABELS["total_distinct_students"], metrics.get("total_distinct_students", 0)))
        rows.append(("", ""))

        rows.append(("◆  CONTACT MATCHING  ◆", ""))
        rows.append((SUMMARY_LABELS["contact_matches"], metrics.get("contact_matches", 0)))
        rows.append((SUMMARY_LABELS["contact_misses"], metrics.get("contact_misses", 0)))
        rows.append(("", ""))

        rows.append(("◆  GROUP ASSIGNMENT  ◆", ""))
        for tab_name in group_order:
            df = group_data.get(tab_name, pd.DataFrame())
            rows.append((f"  {tab_name}", len(df)))
        rows.append(("", ""))

        rows.append(("◆  UNMATCHED BUCKETS  ◆", ""))
        rows.append((
            SUMMARY_LABELS["total_risk_1_2"],
            len(group_data.get(UNMATCHED_LOW_TAB, pd.DataFrame())),
        ))
        rows.append((
            SUMMARY_LABELS["total_risk_3_plus"],
            len(group_data.get(UNMATCHED_HIGH_TAB, pd.DataFrame())),
        ))
        rows.append((
            SUMMARY_LABELS["total_unmatched"],
            metrics.get("total_unmatched", 0),
        ))
        rows.append(("", ""))

        rows.append(("◆  SOURCE FILES  ◆", ""))
        for label, fname in source_files.items():
            rows.append((label, fname))

        for r_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=value)

        apply_summary_formatting(ws)

        # Add visual enhancements
        try:
            enhancer = SummaryEnhancer()
            enhancer.enhance(ws, metrics, group_data, group_order)
        except Exception as exc:
            logger.warning("Exporter: Could not add summary charts: %s", exc)

        logger.info("Exporter: Summary tab written.")

    # ------------------------------------------------------------------
    # Missing Contacts tab
    # ------------------------------------------------------------------

    def _write_missing_contacts(
        self,
        wb: Workbook,
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        used_tab_names: List[str],
    ) -> None:
        """Write a tab listing students with no phone or email found."""
        from utils.config import UNMATCHED_LOW_TAB, UNMATCHED_HIGH_TAB
        from utils.excel_utils import make_header_fill, make_header_font, make_body_font

        all_tabs = list(group_order) + [UNMATCHED_LOW_TAB, UNMATCHED_HIGH_TAB]
        frames = []
        for tab in all_tabs:
            df = group_data.get(tab, pd.DataFrame())
            if df.empty:
                continue
            phone_col = "Phone Number" if "Phone Number" in df.columns else None
            email_col = "Email" if "Email" in df.columns else None
            if phone_col and email_col:
                mask = (df[phone_col].fillna("") == "") & (df[email_col].fillna("") == "")
                missing = df[mask][["Student Name", "Student ID", "Matched Group"]].copy()
                if not missing.empty:
                    frames.append(missing)

        if not frames:
            return  # No missing contacts — skip tab entirely

        missing_df = pd.concat(frames, ignore_index=True).drop_duplicates("Student ID")

        safe = safe_excel_tab_name("Missing_Contacts", used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        header_fill = make_header_fill("843C0C")
        header_font = make_header_font()
        body_font   = make_body_font()

        cols = ["Student Name", "Student ID", "Matched Group"]
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(1, c_idx, col)
            cell.fill = header_fill
            cell.font = header_font

        from openpyxl.styles import Alignment, PatternFill
        from utils.excel_utils import _argb
        alt = PatternFill("solid", fgColor=_argb("DCE6F1"))
        for r_idx, row in enumerate(missing_df.itertuples(index=False), start=2):
            fill = alt if r_idx % 2 == 0 else PatternFill("solid", fgColor=_argb("FFFFFF"))
            for c_idx, val in enumerate([row.Student_Name if hasattr(row,'Student_Name') else "",
                                          row.Student_ID if hasattr(row,'Student_ID') else "",
                                          row.Matched_Group if hasattr(row,'Matched_Group') else ""], start=1):
                cell = ws.cell(r_idx, c_idx, str(val) if val else "")
                cell.fill = fill
                cell.font = body_font

        ws.freeze_panes = ws.cell(2, 1)
        ws.auto_filter.ref = f"A1:C1"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 20

        logger.info("Exporter: Missing_Contacts tab — %d students.", len(missing_df))

    # ------------------------------------------------------------------
    # QA Log tab
    # ------------------------------------------------------------------

    def _write_qa_log(
        self, wb: Workbook, qa_log: QALog, used_tab_names: List[str]
    ) -> None:
        safe = safe_excel_tab_name(QA_LOG_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        entries = qa_log.entries()

        for col_idx, col_name in enumerate(QA_LOG_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, entry in enumerate(entries, start=2):
            for col_idx, col_name in enumerate(QA_LOG_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=entry.get(col_name, ""))

        apply_qa_formatting(ws)
        logger.info("Exporter: QA_Log tab written — %d entries.", len(entries))

    # ------------------------------------------------------------------
    # Processing Manifest tab
    # ------------------------------------------------------------------

    def _write_manifest(
        self,
        wb: Workbook,
        metrics: Dict[str, Any],
        source_files: Dict[str, str],
        used_tab_names: List[str],
    ) -> None:
        safe = safe_excel_tab_name(MANIFEST_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        rows = []
        rows.append(("◆  PROCESSING MANIFEST  ◆", ""))
        rows.append(("", ""))
        rows.append(("Application", APP_NAME))
        rows.append(("Version", APP_VERSION))
        rows.append(("Python Version", sys.version.split()[0]))
        rows.append(("Platform", platform.platform()))
        rows.append(("Processing Timestamp", metrics.get("processing_timestamp", "")))
        rows.append(("Execution Duration (s)", f"{metrics.get('execution_duration', 0):.2f}"))
        rows.append(("", ""))

        rows.append(("◆  SOURCE FILES  ◆", ""))
        for label, fname in source_files.items():
            rows.append((label, fname))
        rows.append(("", ""))

        rows.append(("◆  ROW COUNTS  ◆", ""))
        rows.append(("Total Input Rows", metrics.get("total_input_rows", 0)))
        rows.append(("Total At-Risk Rows", metrics.get("total_at_risk_rows", 0)))
        rows.append(("Duplicate Course Rows Removed", metrics.get("duplicate_course_rows_removed", 0)))
        rows.append(("Distinct At-Risk Students", metrics.get("total_distinct_students", 0)))
        rows.append(("Assigned to Groups", metrics.get("total_assigned", 0)))
        rows.append(("Unmatched", metrics.get("total_unmatched", 0)))
        rows.append(("", ""))

        rows.append(("◆  OUTPUT FILE  ◆", ""))
        rows.append(("Output File", metrics.get("output_filename", "")))

        for r_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=value)

        apply_manifest_formatting(ws)
        logger.info("Exporter: Processing_Manifest tab written.")
