"""
report_status_exporter.py — Builds the faculty completion Excel workbook.

Tabs:
    Overview        — overall stats + bar chart
    By_College      — college table + horizontal bar chart
    By_Department   — department table + horizontal bar chart
    By_Professor    — full professor detail table
    Faculty_Download — one row per faculty: First, Last, Email, College, Dept
"""

import logging
import io
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")   # headless — no display needed
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

from processors.report_status_processor import ReportStatusProcessor
from utils.excel_utils import _argb, make_header_fill, make_header_font, make_body_font

logger = logging.getLogger("intervention_sorter")

# Chart color palette
COLOR_SUBMITTED    = "#2E86AB"   # Blue
COLOR_NOT_SUBMITTED = "#E84855"  # Red
COLOR_BG           = "#F4F6FB"
HEADER_COLOR       = "1F3864"
COLLEGE_HEADER     = "2F5496"
DEPT_HEADER        = "375623"
PROFESSOR_HEADER   = "843C0C"
DOWNLOAD_HEADER    = "4A235A"


class ReportStatusExporter:

    def export(
        self,
        processor: ReportStatusProcessor,
        output_path: Path,
        source_filename: str,
    ) -> None:
        logger.info("ReportStatusExporter: Building workbook → '%s'", output_path.name)

        wb = Workbook()
        wb.remove(wb.active)

        overall = processor.overall_stats()
        df_college = processor.by_college()
        df_dept    = processor.by_department()
        df_prof    = processor.by_professor()
        df_dl      = processor.faculty_download()

        self._write_overview(wb, overall, df_college, source_filename)
        self._write_college_tab(wb, df_college)
        self._write_department_tab(wb, df_dept)
        self._write_professor_tab(wb, df_prof)
        self._write_faculty_download(wb, df_dl)

        try:
            wb.save(output_path)
            logger.info("ReportStatusExporter: Saved → '%s'", output_path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Cannot save — file may be open in Excel.\nPath: {output_path}\n{exc}"
            ) from exc

    # ------------------------------------------------------------------
    # Overview tab
    # ------------------------------------------------------------------

    def _write_overview(
        self,
        wb: Workbook,
        overall: Dict[str, Any],
        df_college: pd.DataFrame,
        source_filename: str,
    ) -> None:
        ws = wb.create_sheet("Overview")

        # Header banner
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "Faculty Progress Report — Completion Overview"
        c.font = Font(name="Calibri", size=14, bold=True, color=_argb("FFFFFF"))
        c.fill = PatternFill("solid", fgColor=_argb(HEADER_COLOR))
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Key stats
        stats = [
            ("", ""),
            ("Source File", source_filename),
            ("Generated", datetime.now().strftime("%m/%d/%Y %I:%M %p")),
            ("", ""),
            ("Total Sections", overall["total_sections"]),
            ("Submitted", overall["submitted"]),
            ("Not Submitted", overall["not_submitted"]),
            ("Overall Completion", f"{overall['completion_pct']}%"),
        ]
        for r_idx, (label, value) in enumerate(stats, start=2):
            ws.cell(r_idx, 1, label).font = Font(name="Calibri", size=11, bold=bool(label))
            ws.cell(r_idx, 2, value).font = Font(name="Calibri", size=11)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 35

        # Overall donut chart
        chart_img = self._make_donut_chart(
            submitted=overall["submitted"],
            not_submitted=overall["not_submitted"],
            pct=overall["completion_pct"],
            title="Overall Submission Rate",
        )
        ws.add_image(chart_img, "D2")

        # College summary table below
        start_row = 18
        ws.cell(start_row, 1).value = "Completion by College"
        ws.cell(start_row, 1).font = Font(name="Calibri", size=11, bold=True,
                                           color=_argb("FFFFFF"))
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor=_argb(COLLEGE_HEADER))

        headers = ["College", "Total Sections", "Submitted", "Not Submitted", "Completion %"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(start_row + 1, c_idx, h)
            cell.font = Font(name="Calibri", size=10, bold=True, color=_argb("FFFFFF"))
            cell.fill = PatternFill("solid", fgColor=_argb(COLLEGE_HEADER))
            cell.alignment = Alignment(horizontal="center")

        alt_fill = PatternFill("solid", fgColor=_argb("DCE6F1"))
        for r_idx, row in df_college.iterrows():
            excel_row = start_row + 2 + r_idx
            fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor=_argb("FFFFFF"))
            values = [row["College"], row["Total Sections"], row["Submitted"],
                      row["Not Submitted"], f"{row['Completion %']}%"]
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(excel_row, c_idx, val)
                cell.fill = fill
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

        for c_idx, w in enumerate([25, 15, 12, 15, 13], 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

    # ------------------------------------------------------------------
    # By College tab
    # ------------------------------------------------------------------

    def _write_college_tab(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("By_College")
        self._write_completion_table(ws, df, COLLEGE_HEADER, "Completion by College")

        chart_img = self._make_bar_chart(
            df, label_col="College",
            title="Submission Rate by College",
            figsize=(7, max(3, len(df) * 0.35 + 1)),
        )
        ws.add_image(chart_img, f"A{len(df) + 5}")

    # ------------------------------------------------------------------
    # By Department tab
    # ------------------------------------------------------------------

    def _write_department_tab(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("By_Department")
        df_display = df.copy()
        self._write_completion_table(ws, df_display, DEPT_HEADER, "Completion by Department",
                                     extra_cols=["College", "Department"])

        chart_img = self._make_bar_chart(
            df, label_col="Department",
            title="Submission Rate by Department",
            figsize=(8, max(4, len(df) * 0.3 + 1)),
        )
        ws.add_image(chart_img, f"A{len(df) + 5}")

    # ------------------------------------------------------------------
    # By Professor tab
    # ------------------------------------------------------------------

    def _write_professor_tab(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("By_Professor")
        cols = ["Last Name", "First Name", "Email", "College", "Department",
                "Total Sections", "Submitted", "Not Submitted", "Completion %"]
        df_out = df[cols].copy()
        df_out["Completion %"] = df_out["Completion %"].apply(lambda x: f"{x}%")
        self._write_table(ws, df_out, PROFESSOR_HEADER, "Faculty Submission Detail")

    # ------------------------------------------------------------------
    # Faculty Download tab
    # ------------------------------------------------------------------

    def _write_faculty_download(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Faculty_Download")
        self._write_table(ws, df, DOWNLOAD_HEADER, "Faculty List")

    # ------------------------------------------------------------------
    # Shared table writer
    # ------------------------------------------------------------------

    def _write_completion_table(
        self,
        ws,
        df: pd.DataFrame,
        header_color: str,
        title: str,
        extra_cols: list = None,
    ) -> None:
        base_cols = ["Total Sections", "Submitted", "Not Submitted", "Completion %"]
        if extra_cols:
            display_cols = extra_cols + base_cols
        else:
            display_cols = [c for c in df.columns if c in base_cols] or list(df.columns)

        df_out = df[[c for c in display_cols if c in df.columns]].copy()
        df_out["Completion %"] = df_out["Completion %"].apply(lambda x: f"{x}%")
        self._write_table(ws, df_out, header_color, title)

    def _write_table(
        self,
        ws,
        df: pd.DataFrame,
        header_color: str,
        title: str,
    ) -> None:
        # Title row
        ws.cell(1, 1, title)
        ws.cell(1, 1).font = Font(name="Calibri", size=12, bold=True,
                                   color=_argb("FFFFFF"))
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=_argb(header_color))
        ws.row_dimensions[1].height = 22

        # Headers
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(2, c_idx, col)
            cell.font = Font(name="Calibri", size=10, bold=True, color=_argb("FFFFFF"))
            cell.fill = PatternFill("solid", fgColor=_argb(header_color))
            cell.alignment = Alignment(horizontal="center" if c_idx > 2 else "left",
                                       vertical="center")

        # Freeze
        ws.freeze_panes = ws.cell(3, 1)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}2"

        alt_fill = PatternFill("solid", fgColor=_argb("DCE6F1"))
        body_font = Font(name="Calibri", size=10)

        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            fill = alt_fill if r_idx % 2 == 1 else PatternFill("solid", fgColor=_argb("FFFFFF"))
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, str(val) if pd.notna(val) else "")
                cell.fill = fill
                cell.font = body_font
                cell.alignment = Alignment(
                    horizontal="center" if c_idx > 2 else "left", vertical="top"
                )

        # Auto column widths
        for c_idx, col_name in enumerate(df.columns, 1):
            max_len = len(col_name)
            for r_idx in range(3, min(ws.max_row + 1, 300)):
                val = ws.cell(r_idx, c_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 50)

    # ------------------------------------------------------------------
    # Chart generators (matplotlib → PNG → openpyxl Image)
    # ------------------------------------------------------------------

    def _chart_to_image(self, fig) -> XLImage:
        """Convert matplotlib figure to openpyxl Image object."""
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=96, bbox_inches="tight",
                    facecolor=COLOR_BG)
        buf.seek(0)
        plt.close(fig)
        img = XLImage(buf)
        img.width  = 480
        img.height = 300
        return img

    def _make_donut_chart(
        self,
        submitted: int,
        not_submitted: int,
        pct: float,
        title: str,
    ) -> XLImage:
        fig, ax = plt.subplots(figsize=(4, 3.2), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)

        sizes  = [submitted, not_submitted] if (submitted + not_submitted) > 0 else [1, 0]
        colors = [COLOR_SUBMITTED, COLOR_NOT_SUBMITTED]
        wedges, _ = ax.pie(
            sizes, colors=colors, startangle=90,
            wedgeprops=dict(width=0.5, edgecolor="white", linewidth=2),
        )
        ax.text(0, 0, f"{pct}%", ha="center", va="center",
                fontsize=22, fontweight="bold", color="#1F3864")
        ax.set_title(title, fontsize=12, fontweight="bold", color="#1F3864", pad=12)

        legend = [
            mpatches.Patch(color=COLOR_SUBMITTED,     label=f"Submitted ({submitted:,})"),
            mpatches.Patch(color=COLOR_NOT_SUBMITTED, label=f"Not Submitted ({not_submitted:,})"),
        ]
        ax.legend(handles=legend, loc="lower center", bbox_to_anchor=(0.5, -0.08),
                  ncol=2, fontsize=9, framealpha=0)
        return self._chart_to_image(fig)

    def _make_bar_chart(
        self,
        df: pd.DataFrame,
        label_col: str,
        title: str,
        figsize: tuple = (10, 6),
    ) -> XLImage:
        labels  = df[label_col].tolist()
        submitted     = df["Submitted"].tolist()
        not_submitted = df["Not Submitted"].tolist()
        pcts    = df["Completion %"].tolist()

        y = np.arange(len(labels))
        height = 0.35

        fig, ax = plt.subplots(figsize=figsize, facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)

        bars1 = ax.barh(y + height/2, submitted,     height, label="Submitted",
                        color=COLOR_SUBMITTED,     edgecolor="white")
        bars2 = ax.barh(y - height/2, not_submitted, height, label="Not Submitted",
                        color=COLOR_NOT_SUBMITTED, edgecolor="white")

        # Percentage labels on the right
        for i, pct in enumerate(pcts):
            ax.text(
                max(submitted[i], not_submitted[i]) + 0.5,
                y[i],
                f"{pct}%",
                va="center", ha="left",
                fontsize=8, fontweight="bold", color="#1F3864",
            )

        ax.set_yticks(y)
        ax.set_yticklabels(labels, fontsize=9)
        ax.set_xlabel("Number of Sections", fontsize=10)
        ax.set_title(title, fontsize=12, fontweight="bold", color="#1F3864", pad=10)
        ax.legend(fontsize=9, framealpha=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.invert_yaxis()

        plt.tight_layout()
        return self._chart_to_image(fig)
